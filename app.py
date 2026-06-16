import os
import re
import unicodedata
from flask import (Flask, render_template, request, redirect, url_for, flash,
                   jsonify)
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
from config import Config
from models import db, AdminUser, Category, SubCategory, Brand, Product, ProductImage, ProductSpec, Presentation
from models import QuoteRequest, QuoteItem, Banner, GalleryCategory, GalleryImage, AmbientCategory, AmbientImage, SiteConfig, PaintColor, BlogCategory, BlogPost, ProductPresentationColor
import time
import threading
import resend
from seed_data import seed as seed_database

app = Flask(__name__)
app.config.from_object(Config)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.instance_path), exist_ok=True)

db.init_app(app)
with app.app_context():
    # Run auto-migrations
    try:
        db.create_all()
        from sqlalchemy import text
        # Add color column to quote_items
        try:
            db.session.execute(text("ALTER TABLE quote_items ADD COLUMN color VARCHAR(100)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add subcategory_id column to products
        try:
            db.session.execute(text("ALTER TABLE products ADD COLUMN subcategory_id INTEGER REFERENCES subcategories(id)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add safety_sheet_url to products
        try:
            db.session.execute(text("ALTER TABLE products ADD COLUMN safety_sheet_url VARCHAR(300)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add catalog_url to products
        try:
            db.session.execute(text("ALTER TABLE products ADD COLUMN catalog_url VARCHAR(300)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add presentation column to quote_items
        try:
            db.session.execute(text("ALTER TABLE quote_items ADD COLUMN presentation VARCHAR(100)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add colors column to ambient_categories
        try:
            db.session.execute(text("ALTER TABLE ambient_categories ADD COLUMN colors VARCHAR(200)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Add area column to ambient_categories
        try:
            db.session.execute(text("ALTER TABLE ambient_categories ADD COLUMN area VARCHAR(50)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
        # Create blog tables if they don't exist
        try:
            db.create_all()
        except Exception:
            pass
        # Add image_url column to categories
        try:
            db.session.execute(text("ALTER TABLE categories ADD COLUMN image_url VARCHAR(255)"))
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception as e:
        print(f"Migration error: {e}")

    # Asegurar que existe al menos un admin
    if not AdminUser.query.filter_by(username='admin').first():
        admin = AdminUser(username='admin')
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()

mail = Mail(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'

if app.config.get('RESEND_API_KEY'):
    resend.api_key = app.config['RESEND_API_KEY']

def send_async_email(app, msg_data):
    with app.app_context():
        try:
            # Try Resend first if configured
            if app.config.get('RESEND_API_KEY'):
                resend.Emails.send({
                    "from": f"Grupo Dewill <onboarding@resend.dev>", # Update this once domain is verified
                    "to": msg_data['recipients'],
                    "subject": msg_data['subject'],
                    "text": msg_data['body']
                })
                print("Email sent via Resend API")
            else:
                # Fallback to Flask-Mail
                msg = Message(
                    msg_data['subject'],
                    recipients=msg_data['recipients'],
                    body=msg_data['body']
                )
                mail.send(msg)
                print("Email sent via SMTP")
        except Exception as e:
            print(f"Async email failed: {e}")

@login_manager.user_loader
def load_user(user_id):
    return AdminUser.query.get(int(user_id))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def save_upload(file, subfolder=''):
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        name, ext = os.path.splitext(filename)
        filename = f"{name}_{int(time.time())}{ext}"
        folder = os.path.join(app.config['UPLOAD_FOLDER'], subfolder)
        os.makedirs(folder, exist_ok=True)
        file.save(os.path.join(folder, filename))
        return f"uploads/{subfolder}/{filename}" if subfolder else f"uploads/{filename}"
    return None

def get_site_config():
    configs = SiteConfig.query.all()
    return {c.key: c.value for c in configs}

@app.context_processor
def inject_site_config():
    return {'site': get_site_config()}

def slugify(value):
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^\w\s-]', '', value).strip().lower()
    return re.sub(r'[-\s]+', '-', value)

# ==================== PUBLIC ====================

@app.route('/')
def index():
    banners = Banner.query.filter_by(is_active=True).order_by(Banner.order).all()
    featured = Product.query.filter_by(is_featured=True, is_active=True).limit(8).all()
    categories = Category.query.order_by(Category.order).all()
    brands = Brand.query.all()
    return render_template('index.html', banners=banners, featured_products=featured,
                           categories=categories, brands=brands)

@app.route('/productos')
def productos():
    categories = Category.query.order_by(Category.order).all()
    subcategories = SubCategory.query.order_by(SubCategory.name).all()
    brands = Brand.query.order_by(Brand.name).all()
    presentations = Presentation.query.order_by(Presentation.name).all()
    
    # Handle multiple selections
    cat_slugs = request.args.getlist('cat')
    if not cat_slugs and request.args.get('cat'):
        cat_slugs = [request.args.get('cat')]
        
    subcat_slugs = request.args.getlist('subcat')
    if not subcat_slugs and request.args.get('subcat'):
        subcat_slugs = [request.args.get('subcat')]
        
    brand_slugs = request.args.getlist('marca')
    if not brand_slugs and request.args.get('marca'):
        brand_slugs = [request.args.get('marca')]
        
    pres_slugs = request.args.getlist('pres')
    if not pres_slugs and request.args.get('pres'):
        pres_slugs = [request.args.get('pres')]
        
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    
    # Selected specs
    selected_specs = {}
    for key in request.args.keys():
        if key.startswith('spec_'):
            spec_name = key[5:]
            selected_specs[spec_name] = request.args.getlist(key)
            
    query = Product.query.filter_by(is_active=True)
    
    if cat_slugs:
        cat_ids = [c.id for c in Category.query.filter(Category.slug.in_(cat_slugs)).all()]
        if cat_ids:
            query = query.filter(Product.category_id.in_(cat_ids))
            
    if subcat_slugs:
        subcat_ids = [s.id for s in SubCategory.query.filter(SubCategory.slug.in_(subcat_slugs)).all()]
        if subcat_ids:
            query = query.filter(Product.subcategory_id.in_(subcat_ids))
            
    if brand_slugs:
        brand_ids = [b.id for b in Brand.query.filter(Brand.slug.in_(brand_slugs)).all()]
        if brand_ids:
            query = query.filter(Product.brand_id.in_(brand_ids))
            
    if pres_slugs:
        query = query.filter(Product.presentations.any(Presentation.slug.in_(pres_slugs)))
            
    if q:
        search_term = ''.join(['_' if c.lower() in 'aeiouáéíóú' else c for c in q])
        query = query.filter(db.or_(
            Product.name.ilike(f'%{search_term}%'), 
            Product.description.ilike(f'%{search_term}%'),
            Product.sku.ilike(f'%{search_term}%'),
            Product.brand.has(Brand.name.ilike(f'%{search_term}%')),
            Product.category.has(Category.name.ilike(f'%{search_term}%'))
        ))
        
    # Filter by specs
    for spec_name, spec_vals in selected_specs.items():
        if spec_vals:
            query = query.filter(Product.specs.any(db.and_(
                ProductSpec.key == spec_name,
                ProductSpec.value.in_(spec_vals)
            )))
            
    prods = query.order_by(Product.name).paginate(page=page, per_page=12, error_out=False)
    
    # Gather distinct specs for sidebar
    exclude_keys = ['Norma', 'Durabilidad', 'Descripción', 'Resistencia', 'Calidad']
    all_keys = db.session.query(ProductSpec.key).distinct().all()
    spec_filters = {}
    for (k,) in all_keys:
        if k and k not in exclude_keys:
            vals = db.session.query(ProductSpec.value).join(Product).filter(
                Product.is_active == True,
                ProductSpec.key == k
            ).distinct().all()
            clean_vals = sorted(list(set([v[0].strip() for v in vals if v[0] and v[0].strip() and len(v[0].strip()) < 50])))
            if clean_vals:
                spec_filters[k] = clean_vals

    return render_template('productos.html', products=prods, categories=categories, brands=brands,
                           subcategories=subcategories, presentations=presentations,
                           current_cats=cat_slugs, current_brands=brand_slugs, 
                           current_subcats=subcat_slugs, current_pres=pres_slugs,
                           search_query=q, spec_filters=spec_filters, selected_specs=selected_specs)

@app.route('/api/productos/search')
def api_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    search_term = ''.join(['_' if c.lower() in 'aeiouáéíóú' else c for c in q])
    results = Product.query.filter(
        Product.is_active == True,
        db.or_(
            Product.name.ilike(f'%{search_term}%'), 
            Product.description.ilike(f'%{search_term}%'),
            Product.sku.ilike(f'%{search_term}%'),
            Product.brand.has(Brand.name.ilike(f'%{search_term}%')),
            Product.category.has(Category.name.ilike(f'%{search_term}%'))
        )
    ).limit(8).all()
    return jsonify([p.to_dict() for p in results])

@app.route('/api/producto/<int:product_id>/colores')
def api_product_colors(product_id):
    """Returns colors available for a product, optionally filtered by presentation_id."""
    pres_id = request.args.get('pres_id', type=int)
    
    if pres_id:
        # Strict filtering by presentation
        links = ProductPresentationColor.query.filter_by(
            product_id=product_id,
            presentation_id=pres_id
        ).all()
        return jsonify([{'name': lk.color_name, 'hex_code': lk.hex_code} for lk in links])
    
    # Fallback if no presentation selected: all colors from specs
    specs = ProductSpec.query.filter(
        ProductSpec.product_id == product_id,
        ProductSpec.key.ilike('color')
    ).all()
    seen = set()
    result = []
    for s in specs:
        if s.value.lower() not in seen:
            pc = PaintColor.query.filter(PaintColor.name.ilike(s.value)).first()
            result.append({'name': s.value, 'hex_code': pc.hex_code if pc else '#cccccc'})
            seen.add(s.value.lower())
    
    # If still empty and it's a paint product, return all PaintColors
    if not result:
        product = Product.query.get(product_id)
        if product and product.category and 'pintura' in product.category.name.lower():
            all_colors = PaintColor.query.filter_by(is_active=True).order_by(PaintColor.name).all()
            result = [{'name': c.name, 'hex_code': c.hex_code} for c in all_colors]
    
    return jsonify(result)



@app.route('/producto/<int:product_id>')
def producto_detalle(product_id):
    product = Product.query.get_or_404(product_id)
    related = Product.query.filter(Product.category_id == product.category_id,
        Product.id != product.id, Product.is_active == True).limit(4).all()
        
    product_color_specs = list(set([spec.value for spec in product.specs if spec.key.lower() == 'color']))
    
    paint_colors = []
    if product_color_specs:
        for cname in product_color_specs:
            pc = PaintColor.query.filter(PaintColor.name.ilike(cname)).first()
            if pc:
                paint_colors.append({'name': pc.name, 'hex_code': pc.hex_code})
            else:
                # Add a fallback hex code for unknown imported colors
                paint_colors.append({'name': cname, 'hex_code': '#cccccc'})
    elif product.category and 'pintura' in product.category.name.lower():
        # Fallback to all active paint colors
        all_colors = PaintColor.query.filter_by(is_active=True).all()
        paint_colors = [{'name': c.name, 'hex_code': c.hex_code} for c in all_colors]
        
    categories = Category.query.order_by(Category.order).all()
    return render_template('producto_detalle.html', product=product, related=related, paint_colors=paint_colors, categories=categories)

@app.route('/api/cotizacion', methods=['POST'])
def api_cotizacion():
    data = request.get_json()
    if not data or not data.get('items'):
        return jsonify({'error': 'No hay productos'}), 400
    quote = QuoteRequest(client_name=data.get('name',''), client_company=data.get('company',''),
        client_email=data.get('email',''), client_phone=data.get('phone',''), message=data.get('message',''))
    db.session.add(quote)
    db.session.flush()
    for item in data['items']:
        db.session.add(QuoteItem(quote_id=quote.id, product_id=item['product_id'], quantity=item.get('quantity',1), color=item.get('color'), presentation=item.get('presentation')))
    db.session.commit()
    try:
        # Obtener configuración de correo
        notify = SiteConfig.get('notify_email', app.config.get('MAIL_NOTIFY_TO'))
        if notify and app.config.get('MAIL_USERNAME'):
            # Construir cuerpo detallado
            items_text = ""
            for item in data['items']:
                prod = Product.query.get(item['product_id'])
                if prod:
                    color_info = f" - Color: {item.get('color')}" if item.get('color') else ""
                    pres_info = f" - Presentación: {item.get('presentation')}" if item.get('presentation') else ""
                    items_text += f"- {prod.name} (SKU: {prod.sku or 'N/A'}){color_info}{pres_info} x {item.get('quantity', 1)}\n"
            
            body = (
                f"Nueva Solicitud de Cotización #{quote.id}\n"
                f"----------------------------------------\n"
                f"CLIENTE:\n"
                f"Nombre: {quote.client_name}\n"
                f"Empresa: {quote.client_company or 'No especificada'}\n"
                f"Email: {quote.client_email}\n"
                f"Teléfono: {quote.client_phone}\n\n"
                f"MENSAJE:\n{quote.message or 'Sin mensaje'}\n\n"
                f"PRODUCTOS:\n{items_text}\n"
                f"----------------------------------------\n"
                f"Ver en el panel: {url_for('admin_quote_detail', qid=quote.id, _external=True)}"
            )
            
            msg_data = {
                'subject': f'Nueva Solicitud de Cotización #{quote.id}',
                'recipients': [notify],
                'body': body
            }
            # Send in background to avoid hanging the UI
            threading.Thread(target=send_async_email, args=(app._get_current_object(), msg_data)).start()
    except Exception as e:
        print(f"Email failed: {e}")
    return jsonify({'success': True, 'quote_id': quote.id})

@app.route('/nosotros')
def nosotros():
    return render_template('nosotros.html')

@app.route('/blog')
def blog():
    cat_slug = request.args.get('cat', '')
    page = request.args.get('page', 1, type=int)
    blog_categories = BlogCategory.query.order_by(BlogCategory.name).all()
    query = BlogPost.query.filter_by(is_published=True)
    if cat_slug:
        bc = BlogCategory.query.filter_by(slug=cat_slug).first()
        if bc:
            query = query.filter_by(category_id=bc.id)
    featured = BlogPost.query.filter_by(is_featured=True, is_published=True).order_by(BlogPost.created_at.desc()).first()
    posts = query.order_by(BlogPost.created_at.desc()).paginate(page=page, per_page=9, error_out=False)
    popular = BlogPost.query.filter_by(is_published=True).order_by(BlogPost.created_at.desc()).limit(5).all()
    return render_template('blog.html', posts=posts, blog_categories=blog_categories,
                           featured=featured, popular=popular, current_cat=cat_slug)

@app.route('/blog/<slug>')
def blog_post(slug):
    post = BlogPost.query.filter_by(slug=slug, is_published=True).first_or_404()
    blog_categories = BlogCategory.query.order_by(BlogCategory.name).all()
    related = BlogPost.query.filter(
        BlogPost.category_id == post.category_id,
        BlogPost.id != post.id,
        BlogPost.is_published == True
    ).limit(3).all()
    return render_template('blog_post.html', post=post, blog_categories=blog_categories, related=related)

@app.route('/contacto', methods=['GET', 'POST'])
def contacto():
    import urllib.parse
    site = SiteConfig.query.first()
    success = False
    whatsapp_url = ""
    if request.method == 'POST':
        nombre = request.form.get('nombre', '')
        email = request.form.get('email', '')
        telefono = request.form.get('telefono', '')
        asunto = request.form.get('asunto', '')
        mensaje = request.form.get('mensaje', '')
        
        # Obtener el correo configurado en el panel administrativo o el de por defecto
        admin_email = app.config['MAIL_NOTIFY_TO']
        if site and site.config_data.get('email_notificaciones'):
            admin_email = site.config_data.get('email_notificaciones')
            
        # Preparar y enviar correo
        msg_data = {
            'recipients': [admin_email],
            'subject': f"Nuevo mensaje de contacto: {asunto}",
            'body': f"Nombre: {nombre}\nEmail: {email}\nTeléfono/WhatsApp: {telefono}\nAsunto: {asunto}\n\nMensaje:\n{mensaje}"
        }
        thread = threading.Thread(target=send_async_email, args=(app._get_current_object(), msg_data))
        thread.start()
        
        # Preparar URL de WhatsApp usando el número configurado en el panel
        wa_number = '51977585654'
        if site and site.config_data.get('whatsapp'):
            wa_number = site.config_data.get('whatsapp')
            
        wa_text = f"Hola Grupo Dewill, mi nombre es {nombre}. {mensaje}"
        whatsapp_url = f"https://wa.me/{wa_number}?text={urllib.parse.quote(wa_text)}"
        
        success = True
    return render_template('contacto.html', success=success, whatsapp_url=whatsapp_url)

@app.route('/galeria')
def galeria():
    gallery_cats = GalleryCategory.query.all()
    current_cat = request.args.get('cat', '')
    if current_cat:
        gc = GalleryCategory.query.filter_by(slug=current_cat).first()
        images = gc.images if gc else []
    else:
        images = GalleryImage.query.order_by(GalleryImage.order).all()
    return render_template('galeria.html', gallery_categories=gallery_cats, images=images, current_cat=current_cat)

@app.route('/galeria/ambientes')
def galeria_ambientes():
    colors = AmbientCategory.query.filter_by(type='color').order_by(AmbientCategory.order).all()
    rooms = AmbientCategory.query.filter_by(type='room').order_by(AmbientCategory.order).all()
    return render_template('galeria_ambientes.html', colors=colors, rooms=rooms)

@app.route('/api/ambientes/<int:id>')
def api_ambiente(id):
    a = AmbientCategory.query.get_or_404(id)
    return jsonify({
        'id': a.id,
        'name': a.name,
        'images': [{'image_url': img.image_url} for img in a.images]
    })


# ==================== ADMIN ROUTES ====================

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        user = AdminUser.query.filter_by(username=request.form.get('username')).first()
        if user and user.check_password(request.form.get('password')):
            login_user(user)
            return redirect(url_for('admin_dashboard'))
        flash('Credenciales inválidas', 'error')
    return render_template('admin/login.html')

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('admin_login'))

@app.route('/admin/')
@login_required
def admin_dashboard():
    total_q = QuoteRequest.query.count()
    responded_q = QuoteRequest.query.filter_by(status='respondida').count()
    processing_q = QuoteRequest.query.filter_by(status='en_proceso').count()
    conversion_rate = 0
    if total_q > 0:
        conversion_rate = round(((responded_q + processing_q) / total_q) * 100, 1)

    stats = {
        'total_products': Product.query.count(),
        'active_products': Product.query.filter_by(is_active=True).count(),
        'total_quotes': total_q,
        'pending_quotes': QuoteRequest.query.filter_by(status='pendiente').count(),
        'responded_quotes': responded_q,
        'processing_quotes': processing_q,
        'conversion_rate': conversion_rate,
        'categories': Category.query.count(),
        'subcategories': SubCategory.query.count(),
        'brands': Brand.query.count(),
        'presentations': Presentation.query.count()
    }
    recent = QuoteRequest.query.order_by(QuoteRequest.created_at.desc()).limit(5).all()
    return render_template('admin/dashboard.html', stats=stats, recent_quotes=recent)

@app.route('/admin/sync-colors', methods=['POST'])
@login_required
def admin_sync_colors():
    # Extended dictionary of common paint colors in Spanish
    hex_map = {
        'blanco': '#ffffff', 'negro': '#000000', 'rojo': '#e60000', 'azul': '#0055a4',
        'verde': '#2e8b57', 'amarillo': '#ffd700', 'gris': '#808080', 'plomo': '#696969',
        'naranja': '#ff8c00', 'celeste': '#87ceeb', 'rosado': '#ffc0cb', 'marron': '#8b4513',
        'marrón': '#8b4513', 'crema': '#fffdd0', 'marfil': '#fffff0', 'arena': '#f4a460',
        'lila': '#c8a2c8', 'morado': '#800080', 'turquesa': '#40e0d0', 'guinda': '#800000',
        'pino': '#01796f', 'almendra': '#efdecd', 'maiz': '#fbec5d', 'durazno': '#ffe5b4',
        'limon': '#fff700', 'limón': '#fff700', 'manzana': '#8db600', 'acero': '#4682b4',
        'coral': '#ff7f50', 'salmon': '#fa8072', 'salmón': '#fa8072', 'vino': '#722f37',
        'mostaza': '#ffdb58', 'cielo': '#87ceeb', 'nieve': '#fffafa', 'hueso': '#e3dac9',
        'perla': '#eae0c8', 'humo': '#738276', 'plata': '#c0c0c0', 'dorado': '#ffd700',
        'oro': '#ffd700', 'cobre': '#b87333', 'bronce': '#cd7f32', 'caoba': '#c04000',
        'pardo': '#5c4033', 'magenta': '#ff00ff', 'cyan': '#00ffff', 'cian': '#00ffff',
        'transparente': '#ffffff', 'ostra': '#e3daca', 'champagne': '#f7e7ce',
        'paja': '#e4d96f', 'tabaco': '#715c3c', 'cemento': '#8c8c8c', 'asfalto': '#595959',
        'ladrillo': '#b22222', 'arcilla': '#b66a50', 'terracota': '#e2725b', 'menta': '#98ff98',
        'olivo': '#808000', 'esmeralda': '#50c878', 'jade': '#00a86b', 'marino': '#000080',
        'cobalto': '#0047ab', 'zafiro': '#0f52ba', 'indigo': '#4b0082', 'índigo': '#4b0082',
        'violeta': '#ee82ee', 'fucsia': '#ff00ff', 'rosa': '#ffc0cb', 'melocoton': '#ffcba4',
        'melocotón': '#ffcba4', 'mandarina': '#f28500', 'caramelo': '#c68e17', 'chocolate': '#7b3f00',
        'cafe': '#6f4e37', 'café': '#6f4e37', 'vainilla': '#f3e5ab', 'almendrado': '#efdecd',
        'marino': '#120a8f', 'noche': '#0c090a', 'pizarra': '#708090', 'tierra': '#a0522d'
    }
    
    # Advanced fallback base words
    base_colors = {
        'blanco': '#f5f5f5', 'nieve': '#fffafa', 'rojo': '#cc0000', 'azul': '#0000cd',
        'verde': '#3cb371', 'amarillo': '#ffda00', 'naranja': '#ff8c00', 'marron': '#8b4513',
        'marrón': '#8b4513', 'gris': '#808080', 'rosa': '#ffb6c1', 'morado': '#8a2be2',
        'crema': '#fffdd0', 'claro': '#e0e0e0', 'oscuro': '#404040'
    }
    
    # Get all unique color names from ProductSpec
    specs = ProductSpec.query.filter(ProductSpec.key.ilike('color')).all()
    unique_colors = set(s.value.strip() for s in specs if s.value)
    
    added = 0
    updated = 0
    
    for cname in unique_colors:
        pc = PaintColor.query.filter(PaintColor.name.ilike(cname)).first()
        
        # Determine hex
        hex_code = '#cccccc'
        lower_name = cname.lower()
        
        if lower_name in hex_map:
            hex_code = hex_map[lower_name]
        else:
            # Try to find exact word in hex_map
            words = lower_name.split()
            found = False
            for w in words:
                if w in hex_map:
                    hex_code = hex_map[w]
                    found = True
                    break
            
            # If still not found, check base colors substring
            if not found:
                for base, hx in base_colors.items():
                    if base in lower_name:
                        hex_code = hx
                        break
            
            # If STILL not found, generate a stable, readable random hex color from the hash of the string
            # to avoid grey boxes! We generate HSL and convert to HEX to ensure it's a nice paint color.
            if hex_code == '#cccccc' and lower_name.strip() != '':
                import hashlib
                h = int(hashlib.md5(cname.encode()).hexdigest(), 16)
                hue = h % 360
                sat = 50 + (h % 30) # 50-80%
                light = 40 + (h % 40) # 40-80%
                
                # Convert HSL to RGB to HEX
                s = sat / 100.0
                l = light / 100.0
                c = (1 - abs(2 * l - 1)) * s
                x = c * (1 - abs((hue / 60.0) % 2 - 1))
                m = l - c / 2.0
                if hue < 60: r,g,b = c,x,0
                elif hue < 120: r,g,b = x,c,0
                elif hue < 180: r,g,b = 0,c,x
                elif hue < 240: r,g,b = 0,x,c
                elif hue < 300: r,g,b = x,0,c
                else: r,g,b = c,0,x
                
                hex_code = '#{:02x}{:02x}{:02x}'.format(int((r+m)*255), int((g+m)*255), int((b+m)*255))
                    
        if not pc:
            pc = PaintColor(name=cname, hex_code=hex_code)
            db.session.add(pc)
            added += 1
        elif pc.hex_code == '#cccccc' or not pc.hex_code or pc.hex_code == '#ffffff' and 'blanco' not in lower_name:
            pc.hex_code = hex_code
            updated += 1
            
        # Update existing ProductPresentationColor links with the correct hex
        links = ProductPresentationColor.query.filter(ProductPresentationColor.color_name.ilike(cname)).all()
        for link in links:
            if link.hex_code != hex_code:
                link.hex_code = hex_code
                    
    db.session.commit()
    flash(f'Colores analizados: {added} nuevos, {updated} actualizados (Hex mejorado). Los vínculos de presentación se han preservado.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/productos')
@login_required
def admin_products():
    q = request.args.get('q', '')
    cat_id = request.args.get('cat', '')
    page = request.args.get('page', 1, type=int)
    query = Product.query
    if q: 
        query = query.filter(db.or_(
            Product.name.ilike(f'%{q}%'), 
            Product.sku.ilike(f'%{q}%'),
            Product.brand.has(Brand.name.ilike(f'%{q}%')),
            Product.category.has(Category.name.ilike(f'%{q}%'))
        ))
    if cat_id: query = query.filter_by(category_id=int(cat_id))
    products = query.order_by(Product.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    categories = Category.query.order_by(Category.order).all()
    return render_template('admin/products.html', products=products, categories=categories, q=q, cat_id=cat_id)

@app.route('/admin/productos/importar', methods=['GET', 'POST'])
@login_required
def admin_product_import():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
        file = request.files['excel_file']
        if file.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(request.url)
            
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=file, read_only=True, data_only=True)
                ws = wb.active
                
                # Read headers from first row
                headers_raw = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                    headers_raw.append(str(cell.value or '').strip().title())
                
                required_cols = ['Codigo', 'Categoria', 'Sub Categoria', 'Marca', 'Producto']
                for col in required_cols:
                    if col not in headers_raw:
                        wb.close()
                        flash(f'Falta la columna requerida: {col}', 'error')
                        return redirect(request.url)
                
                # Build column index map
                col_map = {name: idx for idx, name in enumerate(headers_raw)}
                
                imported = 0
                skipped = 0
                cleared_products = set()
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Safely get cell value by column name
                    def cell(name):
                        idx = col_map.get(name)
                        if idx is None or idx >= len(row):
                            return None
                        v = row[idx]
                        if v is None:
                            return None
                        return str(v).strip() or None
                    
                    codigo = cell('Codigo')
                    if not codigo:
                        skipped += 1
                        continue
                        
                    prod_name = cell('Producto') or 'Sin Nombre'
                    
                    # 1. Handle Brand
                    brand = None
                    marca_name = cell('Marca')
                    if marca_name:
                        marca_slug = slugify(marca_name)
                        brand = Brand.query.filter(Brand.slug == marca_slug).first()
                        if not brand:
                            brand = Brand(name=marca_name, slug=marca_slug)
                            db.session.add(brand)
                            db.session.flush()
                                
                    # 2. Group products by name AND brand
                    if brand:
                        product = Product.query.filter(Product.name.ilike(prod_name), Product.brand_id == brand.id).first()
                    else:
                        product = Product.query.filter(Product.name.ilike(prod_name), Product.brand_id.is_(None)).first()
                    
                    if product and product.id not in cleared_products:
                        ProductPresentationColor.query.filter_by(product_id=product.id).delete()
                        cleared_products.add(product.id)
                        db.session.flush()
                    
                    if not product:
                        if Product.query.filter_by(sku=codigo).first():
                            skipped += 1
                            continue
                        # Handle Category
                        category = None
                        cat_name = cell('Categoria')
                        if cat_name:
                            cat_slug = slugify(cat_name)
                            category = Category.query.filter(Category.slug == cat_slug).first()
                            if not category:
                                category = Category(name=cat_name, slug=cat_slug, order=0)
                                db.session.add(category)
                                db.session.flush()
                                
                        # Handle Sub Category
                        subcategory = None
                        subcat_name = cell('Sub Categoria')
                        if category and subcat_name:
                            subcat_slug = slugify(subcat_name)
                            subcategory = SubCategory.query.filter(SubCategory.slug == subcat_slug, SubCategory.category_id == category.id).first()
                            if not subcategory:
                                subcategory = SubCategory(name=subcat_name, slug=subcat_slug, category_id=category.id)
                                db.session.add(subcategory)
                                db.session.flush()
                                
                        # Create Product
                        product = Product(
                            name=prod_name,
                            sku=codigo,
                            category_id=category.id if category else None,
                            subcategory_id=subcategory.id if subcategory else None,
                            brand_id=brand.id if brand else None,
                            is_active=True,
                            is_featured=False
                        )
                        db.session.add(product)
                        db.session.flush()
                        
                    # Handle Color
                    color_val = cell('Color')
                    if color_val:
                        existing_spec = ProductSpec.query.filter(
                            ProductSpec.product_id == product.id,
                            ProductSpec.key == 'Color',
                            ProductSpec.value.ilike(color_val)
                        ).first()
                        if not existing_spec:
                            db.session.add(ProductSpec(product_id=product.id, key='Color', value=color_val))
                            
                    # Handle Presentation
                    pres = None
                    pres_name = cell('Presentacion')
                    if pres_name:
                        pres_slug = slugify(pres_name)
                        pres = Presentation.query.filter(Presentation.slug == pres_slug).first()
                        if not pres:
                            pres = Presentation(name=pres_name, slug=pres_slug)
                            db.session.add(pres)
                            db.session.flush()
                        if pres not in product.presentations:
                            product.presentations.append(pres)

                    # Link color to presentation
                    if pres and color_val:
                        pc = PaintColor.query.filter(PaintColor.name.ilike(color_val)).first()
                        hex_code = pc.hex_code if pc else '#cccccc'
                        exists_link = ProductPresentationColor.query.filter_by(
                            product_id=product.id,
                            presentation_id=pres.id,
                            color_name=color_val
                        ).first()
                        if not exists_link:
                            db.session.add(ProductPresentationColor(
                                product_id=product.id,
                                presentation_id=pres.id,
                                color_name=color_val,
                                hex_code=hex_code
                            ))

                    imported += 1
                    
                wb.close()
                db.session.commit()
                flash(f'Importación completada. {imported} productos importados, {skipped} omitidos.', 'success')
                return redirect(url_for('admin_products'))
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                try:
                    db.session.rollback()
                except Exception:
                    pass
                error_msg = str(e)
                if len(error_msg) > 200:
                    error_msg = error_msg[:200] + '...'
                flash(f'Error al procesar el archivo Excel: {error_msg}', 'error')
                return redirect(request.url)
        else:
            flash('Formato de archivo no válido. Use .xlsx o .xls', 'error')
            return redirect(request.url)
            
    return render_template('admin/product_import.html')

@app.route('/admin/productos/nuevo', methods=['GET', 'POST'])
@login_required
def admin_product_new():
    categories = Category.query.order_by(Category.order).all()
    subcategories = SubCategory.query.order_by(SubCategory.name).all()
    brands = Brand.query.order_by(Brand.name).all()
    presentations = Presentation.query.order_by(Presentation.name).all()
    if request.method == 'POST':
        p = Product(name=request.form['name'], sku=request.form.get('sku',''),
            description=request.form.get('description',''), technical_description=request.form.get('technical_description',''),
            category_id=request.form.get('category_id',type=int) or None,
            subcategory_id=request.form.get('subcategory_id',type=int) or None,
            brand_id=request.form.get('brand_id',type=int) or None,
            video_url=request.form.get('video_url',''), is_featured='is_featured' in request.form,
            is_active='is_active' in request.form)
        if 'main_image' in request.files and request.files['main_image'].filename:
            p.main_image = save_upload(request.files['main_image'], 'products')
        if 'pdf_file' in request.files and request.files['pdf_file'].filename:
            p.pdf_url = save_upload(request.files['pdf_file'], 'docs')
        if 'safety_sheet_file' in request.files and request.files['safety_sheet_file'].filename:
            p.safety_sheet_url = save_upload(request.files['safety_sheet_file'], 'docs')
        if 'catalog_file' in request.files and request.files['catalog_file'].filename:
            p.catalog_url = save_upload(request.files['catalog_file'], 'docs')
        
        # Save presentations
        pres_ids = request.form.getlist('presentation_ids', type=int)
        for pres_id in pres_ids:
            pres = Presentation.query.get(pres_id)
            if pres: p.presentations.append(pres)

        db.session.add(p)
        db.session.flush()
        for f in request.files.getlist('extra_images'):
            if f.filename:
                url = save_upload(f, 'products')
                if url: db.session.add(ProductImage(product_id=p.id, image_url=url))
        for k, v in zip(request.form.getlist('spec_key[]'), request.form.getlist('spec_value[]')):
            if k.strip() and v.strip():
                db.session.add(ProductSpec(product_id=p.id, key=k.strip(), value=v.strip()))
        db.session.commit()
        flash('Producto creado exitosamente', 'success')
        return redirect(url_for('admin_products'))
    return render_template('admin/product_form.html', product=None, categories=categories, subcategories=subcategories, brands=brands, presentations=presentations)

@app.route('/admin/productos/<int:pid>/editar', methods=['GET', 'POST'])
@login_required
def admin_product_edit(pid):
    p = Product.query.get_or_404(pid)
    categories = Category.query.order_by(Category.order).all()
    subcategories = SubCategory.query.order_by(SubCategory.name).all()
    brands = Brand.query.order_by(Brand.name).all()
    presentations = Presentation.query.order_by(Presentation.name).all()
    if request.method == 'POST':
        p.name = request.form['name']; p.sku = request.form.get('sku','')
        p.description = request.form.get('description',''); p.technical_description = request.form.get('technical_description','')
        p.category_id = request.form.get('category_id',type=int) or None
        p.subcategory_id = request.form.get('subcategory_id',type=int) or None
        p.brand_id = request.form.get('brand_id',type=int) or None
        p.video_url = request.form.get('video_url',''); p.is_featured = 'is_featured' in request.form
        p.is_active = 'is_active' in request.form
        if 'main_image' in request.files and request.files['main_image'].filename:
            p.main_image = save_upload(request.files['main_image'], 'products')
        if 'pdf_file' in request.files and request.files['pdf_file'].filename:
            p.pdf_url = save_upload(request.files['pdf_file'], 'docs')
        if 'safety_sheet_file' in request.files and request.files['safety_sheet_file'].filename:
            p.safety_sheet_url = save_upload(request.files['safety_sheet_file'], 'docs')
        if 'catalog_file' in request.files and request.files['catalog_file'].filename:
            p.catalog_url = save_upload(request.files['catalog_file'], 'docs')
        for f in request.files.getlist('extra_images'):
            if f.filename:
                url = save_upload(f, 'products')
                if url: db.session.add(ProductImage(product_id=p.id, image_url=url))
        ProductSpec.query.filter_by(product_id=p.id).delete()
        for k, v in zip(request.form.getlist('spec_key[]'), request.form.getlist('spec_value[]')):
            if k.strip() and v.strip():
                db.session.add(ProductSpec(product_id=p.id, key=k.strip(), value=v.strip()))
        
        # Update presentations
        p.presentations = []
        pres_ids = request.form.getlist('presentation_ids', type=int)
        for pres_id in pres_ids:
            pres = Presentation.query.get(pres_id)
            if pres: p.presentations.append(pres)
            
        db.session.commit()
        flash('Producto actualizado', 'success')
        return redirect(url_for('admin_products'))
    return render_template('admin/product_form.html', product=p, categories=categories, subcategories=subcategories, brands=brands, presentations=presentations)

@app.route('/admin/productos/<int:pid>/eliminar', methods=['POST'])
@login_required
def admin_product_delete(pid):
    db.session.delete(Product.query.get_or_404(pid)); db.session.commit()
    flash('Producto eliminado', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/productos/eliminar-seleccion', methods=['POST'])
@login_required
def admin_products_bulk_delete():
    try:
        product_ids = request.form.get('product_ids', '')
        if not product_ids:
            flash('No se seleccionaron productos para eliminar.', 'warning')
            return redirect(url_for('admin_products'))
            
        id_list = [int(id_str) for id_str in product_ids.split(',') if id_str.strip().isdigit()]
        if not id_list:
            flash('IDs de productos no válidos.', 'error')
            return redirect(url_for('admin_products'))
            
        products = Product.query.filter(Product.id.in_(id_list)).all()
        count = len(products)
        for p in products:
            db.session.delete(p)
        db.session.commit()
        flash(f'{count} producto(s) han sido eliminados correctamente.', 'success')
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            db.session.rollback()
        except Exception:
            pass
        error_msg = str(e)
        if len(error_msg) > 500:
            error_msg = error_msg[:500] + '...'
        flash(f'Error al eliminar productos: {error_msg}', 'error')
    return redirect(url_for('admin_products'))

@app.route('/admin/productos/imagen/<int:img_id>/eliminar', methods=['POST'])
@login_required
def admin_product_image_delete(img_id):
    img = ProductImage.query.get_or_404(img_id); pid = img.product_id
    db.session.delete(img); db.session.commit()
    return redirect(url_for('admin_product_edit', pid=pid))

@app.route('/admin/categorias', methods=['GET', 'POST'])
@login_required
def admin_categories():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            c = Category(name=request.form['name'], slug=request.form['slug'],
                icon=request.form.get('icon','fas fa-tag'), order=request.form.get('order',0,type=int))
            if 'image' in request.files and request.files['image'].filename:
                c.image_url = save_upload(request.files['image'], 'categories')
            db.session.add(c)
            db.session.commit(); flash('Categoría creada', 'success')
        elif action == 'update':
            c = Category.query.get(request.form['id'])
            if c:
                c.name=request.form['name']; c.slug=request.form['slug']
                c.icon=request.form.get('icon','fas fa-tag'); c.order=request.form.get('order',0,type=int)
                if 'image' in request.files and request.files['image'].filename:
                    c.image_url = save_upload(request.files['image'], 'categories')
                db.session.commit(); flash('Categoría actualizada', 'success')
        elif action == 'delete':
            c = Category.query.get(request.form['id'])
            if c and not c.products:
                db.session.delete(c); db.session.commit(); flash('Categoría eliminada', 'success')
            else: flash('No se puede eliminar: tiene productos', 'error')
        return redirect(url_for('admin_categories'))
    q = request.args.get('q', '')
    query = Category.query
    if q:
        query = query.filter(Category.name.ilike(f'%{q}%'))
    categories = query.order_by(Category.order).all()
    return render_template('admin/categories.html', categories=categories, q=q)

@app.route('/admin/marcas', methods=['GET', 'POST'])
@login_required
def admin_brands():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            b = Brand(name=request.form['name'], slug=request.form['slug'], description=request.form.get('description',''))
            if 'logo' in request.files and request.files['logo'].filename:
                b.logo_url = save_upload(request.files['logo'], 'brands')
            db.session.add(b); db.session.commit(); flash('Marca creada', 'success')
        elif action == 'update':
            b = Brand.query.get(request.form['id'])
            if b:
                b.name=request.form['name']; b.slug=request.form['slug']; b.description=request.form.get('description','')
                if 'logo' in request.files and request.files['logo'].filename:
                    b.logo_url = save_upload(request.files['logo'], 'brands')
                db.session.commit(); flash('Marca actualizada', 'success')
        elif action == 'delete':
            b = Brand.query.get(request.form['id'])
            if b and not b.products:
                db.session.delete(b); db.session.commit(); flash('Marca eliminada', 'success')
            else: flash('No se puede eliminar: tiene productos', 'error')
        return redirect(url_for('admin_brands'))
    q = request.args.get('q', '')
    query = Brand.query
    if q:
        query = query.filter(Brand.name.ilike(f'%{q}%'))
    brands = query.order_by(Brand.name).all()
    return render_template('admin/brands.html', brands=brands, q=q)


@app.route('/admin/subcategorias', methods=['GET', 'POST'])
@login_required
def admin_subcategories():
    categories = Category.query.order_by(Category.order).all()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            name = request.form['name']
            slug = request.form.get('slug') or slugify(name)
            cat_id = request.form.get('category_id', type=int)
            db.session.add(SubCategory(name=name, slug=slug, category_id=cat_id))
            db.session.commit()
            flash('Sub Categoría creada', 'success')
        elif action == 'update':
            sc = SubCategory.query.get(request.form['id'])
            if sc:
                sc.name = request.form['name']
                sc.slug = request.form.get('slug') or slugify(sc.name)
                sc.category_id = request.form.get('category_id', type=int)
                db.session.commit()
                flash('Sub Categoría actualizada', 'success')
        elif action == 'delete':
            sc = SubCategory.query.get(request.form['id'])
            if sc and not sc.products:
                db.session.delete(sc)
                db.session.commit()
                flash('Sub Categoría eliminada', 'success')
            else:
                flash('No se puede eliminar: tiene productos asociados', 'error')
        return redirect(url_for('admin_subcategories'))
    q = request.args.get('q', '')
    query = SubCategory.query.join(Category)
    if q:
        query = query.filter(SubCategory.name.ilike(f'%{q}%'))
    subcategories = query.order_by(Category.order, SubCategory.name).all()
    return render_template('admin/subcategories.html', subcategories=subcategories, categories=categories, q=q)


@app.route('/admin/presentaciones', methods=['GET', 'POST'])
@login_required
def admin_presentations():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            name = request.form['name']
            slug = request.form.get('slug') or slugify(name)
            db.session.add(Presentation(name=name, slug=slug))
            db.session.commit()
            flash('Presentación creada', 'success')
        elif action == 'update':
            p = Presentation.query.get(request.form['id'])
            if p:
                p.name = request.form['name']
                p.slug = request.form.get('slug') or slugify(p.name)
                db.session.commit()
                flash('Presentación actualizada', 'success')
        elif action == 'delete':
            p = Presentation.query.get(request.form['id'])
            if p and not p.products:
                db.session.delete(p)
                db.session.commit()
                flash('Presentación eliminada', 'success')
            else:
                flash('No se puede eliminar: está asociada a productos', 'error')
        return redirect(url_for('admin_presentations'))
    q = request.args.get('q', '')
    query = Presentation.query
    if q:
        query = query.filter(Presentation.name.ilike(f'%{q}%'))
    presentations = query.order_by(Presentation.name).all()
    return render_template('admin/presentations.html', presentations=presentations, q=q)

@app.route('/admin/solicitudes')
@login_required
def admin_quotes():
    status = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    query = QuoteRequest.query
    if status: query = query.filter_by(status=status)
    quotes = query.order_by(QuoteRequest.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/quotes.html', quotes=quotes, current_status=status)

@app.route('/admin/solicitudes/<int:qid>')
@login_required
def admin_quote_detail(qid):
    return render_template('admin/quote_detail.html', quote=QuoteRequest.query.get_or_404(qid))

@app.route('/admin/solicitudes/<int:qid>/estado', methods=['POST'])
@login_required
def admin_quote_status(qid):
    q = QuoteRequest.query.get_or_404(qid)
    q.status = request.form.get('status', 'pendiente')
    db.session.commit(); flash('Estado actualizado', 'success')
    return redirect(url_for('admin_quote_detail', qid=qid))

@app.route('/admin/banners', methods=['GET', 'POST'])
@login_required
def admin_banners():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            b = Banner(title=request.form.get('title',''), subtitle=request.form.get('subtitle',''),
                cta_text=request.form.get('cta_text',''), cta_link=request.form.get('cta_link',''),
                order=request.form.get('order',0,type=int), is_active='is_active' in request.form)
            if 'image' in request.files and request.files['image'].filename:
                b.image_url = save_upload(request.files['image'], 'banners')
            db.session.add(b); db.session.commit(); flash('Banner creado', 'success')
        elif action == 'update':
            b = Banner.query.get(request.form['id'])
            if b:
                b.title=request.form.get('title',''); b.subtitle=request.form.get('subtitle','')
                b.cta_text=request.form.get('cta_text',''); b.cta_link=request.form.get('cta_link','')
                b.order=request.form.get('order',0,type=int); b.is_active='is_active' in request.form
                if 'image' in request.files and request.files['image'].filename:
                    b.image_url = save_upload(request.files['image'], 'banners')
                db.session.commit(); flash('Banner actualizado', 'success')
        elif action == 'delete':
            b = Banner.query.get(request.form['id'])
            if b: db.session.delete(b); db.session.commit(); flash('Banner eliminado', 'success')
        return redirect(url_for('admin_banners'))
    return render_template('admin/banners.html', banners=Banner.query.order_by(Banner.order).all())

@app.route('/admin/galeria', methods=['GET', 'POST'])
@login_required
def admin_gallery():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create_category':
            db.session.add(GalleryCategory(name=request.form['name'], slug=request.form['slug']))
            db.session.commit(); flash('Categoría creada', 'success')
        elif action == 'delete_category':
            gc = GalleryCategory.query.get(request.form['id'])
            if gc: db.session.delete(gc); db.session.commit(); flash('Categoría eliminada', 'success')
        elif action == 'add_image':
            if 'image' in request.files and request.files['image'].filename:
                url = save_upload(request.files['image'], 'gallery')
                if url:
                    db.session.add(GalleryImage(category_id=request.form.get('category_id',type=int),
                        image_url=url, title=request.form.get('title',''),
                        color_code=request.form.get('color_code',''), order=request.form.get('order',0,type=int)))
                    db.session.commit(); flash('Imagen agregada', 'success')
        elif action == 'delete_image':
            gi = GalleryImage.query.get(request.form['id'])
            if gi: db.session.delete(gi); db.session.commit(); flash('Imagen eliminada', 'success')
        return redirect(url_for('admin_gallery'))
    return render_template('admin/gallery.html', gallery_categories=GalleryCategory.query.all(),
        all_images=GalleryImage.query.order_by(GalleryImage.order).all())

@app.route('/admin/configuracion', methods=['GET', 'POST'])
@login_required
def admin_settings():
    if request.method == 'POST':
        for key in ['phone','whatsapp','whatsapp_2','whatsapp_3','email','notify_email','address','facebook','instagram','tiktok',
                     'hero_title','hero_subtitle','hero_badge','meta_title','meta_description','meta_keywords','maps_embed']:
            SiteConfig.set(key, request.form.get(key, ''))
        flash('Configuración guardada', 'success')
        return redirect(url_for('admin_settings'))
    return render_template('admin/settings.html')


@app.route('/admin/blog')
@login_required
def admin_blog():
    page = request.args.get('page', 1, type=int)
    posts = BlogPost.query.order_by(BlogPost.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/blog.html', posts=posts)

@app.route('/admin/blog/nuevo', methods=['GET', 'POST'])
@login_required
def admin_blog_new():
    categories = BlogCategory.query.order_by(BlogCategory.name).all()
    if request.method == 'POST':
        title = request.form['title']
        slug_val = request.form.get('slug') or slugify(title)
        # Ensure unique slug
        base_slug = slug_val
        counter = 1
        while BlogPost.query.filter_by(slug=slug_val).first():
            slug_val = f"{base_slug}-{counter}"
            counter += 1
        post = BlogPost(
            title=title,
            slug=slug_val,
            excerpt=request.form.get('excerpt', ''),
            content=request.form.get('content', ''),
            category_id=request.form.get('category_id', type=int) or None,
            read_time=request.form.get('read_time', 5, type=int),
            is_featured='is_featured' in request.form,
            is_published='is_published' in request.form
        )
        if 'cover_image' in request.files and request.files['cover_image'].filename:
            post.cover_image = save_upload(request.files['cover_image'], 'blog')
        db.session.add(post)
        db.session.commit()
        flash('Artículo creado exitosamente', 'success')
        return redirect(url_for('admin_blog'))
    return render_template('admin/blog_form.html', post=None, categories=categories)

@app.route('/admin/blog/<int:pid>/editar', methods=['GET', 'POST'])
@login_required
def admin_blog_edit(pid):
    post = BlogPost.query.get_or_404(pid)
    categories = BlogCategory.query.order_by(BlogCategory.name).all()
    if request.method == 'POST':
        post.title = request.form['title']
        post.slug = request.form.get('slug') or slugify(post.title)
        post.excerpt = request.form.get('excerpt', '')
        post.content = request.form.get('content', '')
        post.category_id = request.form.get('category_id', type=int) or None
        post.read_time = request.form.get('read_time', 5, type=int)
        post.is_featured = 'is_featured' in request.form
        post.is_published = 'is_published' in request.form
        if 'cover_image' in request.files and request.files['cover_image'].filename:
            post.cover_image = save_upload(request.files['cover_image'], 'blog')
        db.session.commit()
        flash('Artículo actualizado', 'success')
        return redirect(url_for('admin_blog'))
    return render_template('admin/blog_form.html', post=post, categories=categories)

@app.route('/admin/blog/<int:pid>/eliminar', methods=['POST'])
@login_required
def admin_blog_delete(pid):
    db.session.delete(BlogPost.query.get_or_404(pid))
    db.session.commit()
    flash('Artículo eliminado', 'success')
    return redirect(url_for('admin_blog'))

@app.route('/admin/blog-categorias', methods=['GET', 'POST'])
@login_required
def admin_blog_categories():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            name = request.form['name']
            slug = request.form.get('slug') or slugify(name)
            db.session.add(BlogCategory(name=name, slug=slug, icon=request.form.get('icon', 'fas fa-tag')))
            db.session.commit()
            flash('Categoría de blog creada', 'success')
        elif action == 'update':
            bc = BlogCategory.query.get(request.form['id'])
            if bc:
                bc.name = request.form['name']
                bc.slug = request.form.get('slug') or slugify(bc.name)
                bc.icon = request.form.get('icon', 'fas fa-tag')
                db.session.commit()
                flash('Categoría actualizada', 'success')
        elif action == 'delete':
            bc = BlogCategory.query.get(request.form['id'])
            if bc and not bc.posts:
                db.session.delete(bc)
                db.session.commit()
                flash('Categoría eliminada', 'success')
            else:
                flash('No se puede eliminar: tiene artículos asociados', 'error')
        return redirect(url_for('admin_blog_categories'))
    categories = BlogCategory.query.order_by(BlogCategory.name).all()
    return render_template('admin/blog_categories.html', categories=categories)

@app.route('/admin/seed-db', methods=['POST'])
@login_required
def admin_seed_db():
    try:
        seed_database()
        flash('Datos de demostración cargados exitosamente', 'success')
    except Exception as e:
        flash(f'Error al cargar datos: {e}', 'error')
    return redirect(url_for('admin_settings'))


@app.route('/admin/limpiar-duplicados', methods=['POST'])
@login_required
def admin_clean_duplicates():
    """Merge duplicate products (same name) into one, consolidating colors."""
    try:
        products = Product.query.order_by(Product.id).all()

        # Group products by normalized name
        name_map = {}
        for p in products:
            key = p.name.strip().lower()
            if key not in name_map:
                name_map[key] = []
            name_map[key].append(p)

        deleted_count = 0
        merged_colors = 0

        for key, prod_list in name_map.items():
            if len(prod_list) <= 1:
                continue

            base_prod = prod_list[0]
            duplicates = prod_list[1:]

            base_colors = [s.value.lower() for s in base_prod.specs if s.key.lower() == 'color']
            base_pres_ids = [pr.id for pr in base_prod.presentations]

            for dup in duplicates:
                # Merge colors
                for spec in list(dup.specs):
                    if spec.key.lower() == 'color' and spec.value.lower() not in base_colors:
                        db.session.add(ProductSpec(product_id=base_prod.id, key='Color', value=spec.value))
                        base_colors.append(spec.value.lower())
                        merged_colors += 1

                # Merge presentations
                for pres in list(dup.presentations):
                    if pres.id not in base_pres_ids:
                        base_prod.presentations.append(pres)
                        base_pres_ids.append(pres.id)

                # Merge main image if base doesn't have one
                if not base_prod.main_image and dup.main_image:
                    base_prod.main_image = dup.main_image

                db.session.delete(dup)
                deleted_count += 1

        db.session.commit()
        flash(f'Limpieza completada: {deleted_count} productos duplicados eliminados, {merged_colors} colores combinados.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error durante la limpieza: {str(e)}', 'error')
    return redirect(url_for('admin_settings'))


@app.route('/admin/ambientes')
@login_required
def admin_ambientes():
    ambientes = AmbientCategory.query.order_by(AmbientCategory.type, AmbientCategory.order).all()
    return render_template('admin/ambientes.html', ambientes=ambientes)


@app.route('/admin/ambientes/nuevo', methods=['GET', 'POST'])
@login_required
def admin_ambient_new():
    if request.method == 'POST':
        name = request.form['name']
        a = AmbientCategory(
            name=name,
            slug=slugify(name),
            type=request.form['type'],
            description=request.form.get('description', ''),
            colors=request.form.get('colors', ''),
            area=request.form.get('area', 'interior'),
            order=int(request.form.get('order', 0))
        )
        if 'cover_image' in request.files and request.files['cover_image'].filename:
            a.cover_image = save_upload(request.files['cover_image'], 'ambientes')
        
        db.session.add(a)
        db.session.flush()

        for f in request.files.getlist('extra_images'):
            if f.filename:
                url = save_upload(f, 'ambientes')
                if url: db.session.add(AmbientImage(category_id=a.id, image_url=url))
        
        db.session.commit()
        flash('Ambiente creado exitosamente')
        return redirect(url_for('admin_ambientes'))
    return render_template('admin/ambient_form.html', ambient=None)


@app.route('/admin/ambientes/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def admin_ambient_edit(id):
    a = AmbientCategory.query.get_or_404(id)
    if request.method == 'POST':
        a.name = request.form['name']
        a.slug = slugify(a.name)
        a.type = request.form['type']
        a.description = request.form.get('description', '')
        a.colors = request.form.get('colors', '')
        a.area = request.form.get('area', 'interior')
        a.order = int(request.form.get('order', 0))

        if 'cover_image' in request.files and request.files['cover_image'].filename:
            a.cover_image = save_upload(request.files['cover_image'], 'ambientes')

        for f in request.files.getlist('extra_images'):
            if f.filename:
                url = save_upload(f, 'ambientes')
                if url: db.session.add(AmbientImage(category_id=a.id, image_url=url))

        db.session.commit()
        flash('Ambiente actualizado exitosamente')
        return redirect(url_for('admin_ambientes'))
    return render_template('admin/ambient_form.html', ambient=a)

@app.route('/admin/ambientes/<int:id>/eliminar_portada', methods=['POST'])
@login_required
def admin_ambient_delete_cover(id):
    a = AmbientCategory.query.get_or_404(id)
    a.cover_image = None
    db.session.commit()
    flash('Imagen de portada eliminada')
    return redirect(url_for('admin_ambient_edit', id=id))


@app.route('/admin/ambientes/<int:id>/eliminar', methods=['POST'])
@login_required
def admin_ambient_delete(id):
    a = AmbientCategory.query.get_or_404(id)
    db.session.delete(a)
    db.session.commit()
    flash('Ambiente eliminado')
    return redirect(url_for('admin_ambientes'))


@app.route('/admin/ambientes/imagen/<int:img_id>/eliminar', methods=['POST'])
@login_required
def admin_ambient_image_delete(img_id):
    img = AmbientImage.query.get_or_404(img_id)
    parent_id = img.category_id
    db.session.delete(img)
    db.session.commit()
    flash('Imagen eliminada')
    return redirect(url_for('admin_ambient_edit', id=parent_id))


@app.route('/admin/colores', methods=['GET', 'POST'])
@login_required
def admin_colors():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'create':
            db.session.add(PaintColor(name=request.form['name'], hex_code=request.form['hex_code'], is_active='is_active' in request.form))
            db.session.commit(); flash('Color creado', 'success')
        elif action == 'update':
            c = PaintColor.query.get(request.form['id'])
            if c:
                c.name = request.form['name']; c.hex_code = request.form['hex_code']; c.is_active = 'is_active' in request.form
                db.session.commit(); flash('Color actualizado', 'success')
        elif action == 'delete':
            c = PaintColor.query.get(request.form['id'])
            if c: db.session.delete(c); db.session.commit(); flash('Color eliminado', 'success')
        return redirect(url_for('admin_colors'))
    q = request.args.get('q', '')
    query = PaintColor.query
    if q:
        query = query.filter(PaintColor.name.ilike(f'%{q}%') | PaintColor.hex_code.ilike(f'%{q}%'))
    colors = query.all()
    return render_template('admin/colors.html', colors=colors, q=q)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
